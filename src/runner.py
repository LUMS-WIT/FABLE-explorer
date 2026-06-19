#!/usr/bin/env python
"""
Run every pathway in a FABLE workbook, recalculate in Excel, and export outputs.

For each pathway: set the selection cell → trigger Excel recalculation → read
output table values directly from the live workbook → write CSVs.

Aggregated outputs:
- Combined CSV per output table with a RunPathway column
- Scenario deviation summary vs baseline
- Run manifest with status per pathway
"""

from __future__ import annotations

import argparse
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

from comparison import export_scenario_deviation_summary, safe_name

try:
    import xlwings as xw
except ImportError as exc:  # pragma: no cover
    raise SystemExit("xlwings is required. Install with: pip install xlwings") from exc


TABLE_OUTPUT_SHEETS = [
    "GHG",
    "PRODUCTION",
    "TRADE",
    "JOBS",
    "FOOD",
    "LAND",
    "WATER",
    "N and P",
    "BIODIVERSITY",
]
DEVIATION_SUMMARY_FILE = "scenario_deviation_summary.csv"


@dataclass
class PathwaySelectionInfo:
    sheet_name: str
    selection_col_idx: int
    pathway_col_idx: int
    first_data_row: int
    last_data_row: int
    pathways: List[Tuple[int, str]]


RUN_PATHWAY_COL = "RunPathway"


def dedupe_headers(headers: List[object]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for i, header in enumerate(headers, start=1):
        base = str(header).strip() if header is not None else f"col_{i}"
        if not base:
            base = f"col_{i}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        out.append(base if count == 0 else f"{base}_{count+1}")
    return out


def table_ref(table_obj) -> str:
    if isinstance(table_obj, str):
        return table_obj
    return table_obj.ref


def _bounds(ref: str) -> Tuple[int, int, int, int]:
    b = range_boundaries(ref)
    return (b[0] or 0), (b[1] or 0), (b[2] or 0), (b[3] or 0)


def get_pathway_selection_info(workbook_path: Path) -> PathwaySelectionInfo:
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws = wb["PATHWAYS selection"]
        if "PathwaysSelection" not in ws.tables:
            raise ValueError("Table 'PathwaysSelection' not found in PATHWAYS selection.")
        psel_ref = table_ref(ws.tables["PathwaysSelection"])
        min_col, min_row, max_col, max_row = _bounds(psel_ref)
        headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
        normalized = [str(h).strip().lower() if h is not None else "" for h in headers]

        if "selection" not in normalized or "pathway" not in normalized:
            raise ValueError("PathwaysSelection table must include 'SELECTION' and 'PATHWAY' headers.")

        selection_col_idx = min_col + normalized.index("selection")
        pathway_col_idx = min_col + normalized.index("pathway")

        first_data_row = min_row + 1
        last_data_row = max_row
        pathways: List[Tuple[int, str]] = []
        for row in range(first_data_row, last_data_row + 1):
            pathway_name = ws.cell(row, pathway_col_idx).value
            if pathway_name is None or str(pathway_name).strip() == "":
                continue
            pathways.append((row, str(pathway_name).strip()))

        if not pathways:
            raise ValueError("No pathway rows found in PathwaysSelection table.")

        return PathwaySelectionInfo(
            sheet_name="PATHWAYS selection",
            selection_col_idx=selection_col_idx,
            pathway_col_idx=pathway_col_idx,
            first_data_row=first_data_row,
            last_data_row=last_data_row,
            pathways=pathways,
        )
    finally:
        wb.close()


def discover_table_ranges(workbook_path: Path) -> Dict[Tuple[str, str], str]:
    """Pre-scan once to find named Excel table ranges on all output sheets."""
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    ranges: Dict[Tuple[str, str], str] = {}
    try:
        for sheet_name in TABLE_OUTPUT_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for table_name, table_obj in ws.tables.items():
                ranges[(sheet_name, table_name)] = table_ref(table_obj)
    finally:
        wb.close()
    return ranges


def _read_table_from_xw(
    com_wb: "xw.Book",
    sheet_name: str,
    ref: str,
    pathway_name: str,
    out_file: Path,
) -> Optional[pd.DataFrame]:
    """Read a table range from the live Excel workbook via xlwings and write to CSV."""
    try:
        raw = com_wb.sheets[sheet_name].range(ref).value
    except Exception:
        return None
    if not raw:
        return None
    # xlwings returns a flat list for single-row ranges; normalise to list-of-lists
    if not isinstance(raw[0], (list, tuple)):
        raw = [raw]
    headers = dedupe_headers(raw[0] or [])
    df = pd.DataFrame(raw[1:], columns=headers).dropna(how="all")
    if df.empty:
        return None
    pathway_col = RUN_PATHWAY_COL
    if pathway_col in df.columns:
        pathway_col = f"_{RUN_PATHWAY_COL}"
    df.insert(0, pathway_col, pathway_name)
    out_file.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_file, index=False)
    return df


def export_combined_tables(
    combined_tables: Dict[Tuple[str, str], List[pd.DataFrame]],
    combined_dir: Path,
) -> None:
    combined_dir.mkdir(parents=True, exist_ok=True)
    for (sheet_name, table_name), frames in combined_tables.items():
        if not frames:
            continue
        all_df = pd.concat(frames, ignore_index=True)
        out_name = f"{safe_name(sheet_name)}__{safe_name(table_name)}__all_pathways.csv"
        all_df.to_csv(combined_dir / out_name, index=False)


def run_all_pathways(
    workbook_path: Path,
    output_root: Path,
    max_pathways: Optional[int] = None,
    excel_visible: bool = False,
    progress_callback: Optional[Callable[[int, int, str, str], None]] = None,
) -> Path:
    print(f"Workbook : {workbook_path}", flush=True)
    print(f"Outputs  : {output_root}", flush=True)

    selection = get_pathway_selection_info(workbook_path)
    pathways = selection.pathways[:max_pathways] if max_pathways else selection.pathways

    table_ranges = discover_table_ranges(workbook_path)
    if table_ranges:
        print(f"Tables   : {len(table_ranges)} found across output sheets", flush=True)
    else:
        print("Warning  : no named Excel tables found on output sheets — CSVs will be empty.", flush=True)

    run_dir = output_root / f"all_pathways_run_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    tables_dir = run_dir / "tables_per_pathway"
    combined_tables_dir = run_dir / "combined_tables"
    run_dir.mkdir(parents=True, exist_ok=True)

    combined_tables: Dict[Tuple[str, str], List[pd.DataFrame]] = {}
    manifest_rows: List[dict] = []

    app = xw.App(visible=bool(excel_visible), add_book=False)
    app.display_alerts = False
    com_wb = None
    try:
        com_wb = app.books.open(str(workbook_path.resolve()), update_links=False)
        sheet = com_wb.sheets[selection.sheet_name]
        selection_col_letter = get_column_letter(selection.selection_col_idx)
        clear_range = (
            f"{selection_col_letter}{selection.first_data_row}:"
            f"{selection_col_letter}{selection.last_data_row}"
        )

        for i, (row_num, pathway_name) in enumerate(pathways, start=1):
            print(f"[{i}/{len(pathways)}] {pathway_name}", flush=True)
            if progress_callback is not None:
                try:
                    progress_callback(i, len(pathways), pathway_name, "running")
                except Exception:
                    pass
            start = time.time()
            entry: dict = {"Pathway": pathway_name, "Row": row_num, "Status": "ok", "Error": ""}
            try:
                sheet.range(clear_range).value = None
                sheet.cells(row_num, selection.selection_col_idx).value = "x"
                app.calculate()

                tables_found = 0
                for (sht_name, tbl_name), ref in table_ranges.items():
                    out_file = (
                        tables_dir
                        / safe_name(pathway_name)
                        / f"{safe_name(sht_name)}__{safe_name(tbl_name)}.csv"
                    )
                    df = _read_table_from_xw(com_wb, sht_name, ref, pathway_name, out_file)
                    if df is not None:
                        combined_tables.setdefault((sht_name, tbl_name), []).append(df)
                        tables_found += 1
                print(f"         → {tables_found} table(s) extracted", flush=True)

            except Exception as exc:
                entry["Status"] = "failed"
                entry["Error"] = str(exc)
            finally:
                entry["Seconds"] = round(time.time() - start, 2)
                manifest_rows.append(entry)
    finally:
        if com_wb is not None:
            com_wb.close()
        app.quit()

    export_combined_tables(combined_tables, combined_tables_dir)

    deviation_csv = run_dir / DEVIATION_SUMMARY_FILE
    try:
        deviation_df = export_scenario_deviation_summary(
            combined_dir=combined_tables_dir,
            output_csv=deviation_csv,
        )
        print(f"Deviation summary: {deviation_csv} ({len(deviation_df)} rows)", flush=True)
    except Exception as exc:
        print(f"Warning: deviation summary failed: {exc}", flush=True)

    manifest = pd.DataFrame(manifest_rows)
    manifest.to_csv(run_dir / "run_manifest.csv", index=False)
    if progress_callback is not None:
        try:
            progress_callback(len(pathways), len(pathways), "Done", "done")
        except Exception:
            pass

    print(f"Done → {run_dir}", flush=True)
    return run_dir


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run all FABLE pathways and export CSVs.")
    parser.add_argument(
        "--workbook", default=None,
        help="Path to the FABLE workbook. Defaults to config.yaml value.",
    )
    parser.add_argument(
        "--output-dir", default=None,
        help="Output root directory. Default: exports/ in the repo root.",
    )
    parser.add_argument(
        "--max-pathways", type=int, default=None,
        help="Limit to first N pathways (for testing).",
    )
    parser.add_argument(
        "--excel-visible", action="store_true",
        help="Show the Excel window while running.",
    )
    return parser.parse_args()


def _workbook_from_config() -> str:
    import re
    config = Path(__file__).parents[1] / "config.yaml"
    if config.exists():
        for line in config.read_text().splitlines():
            m = re.match(r"^\s*workbook\s*:\s*(.+)", line)
            if m:
                return m.group(1).strip()
    raise SystemExit("No --workbook given and config.yaml has no workbook entry.")


def main() -> None:
    args = parse_args()
    wb = args.workbook or _workbook_from_config()
    workbook_path = Path(wb).expanduser().resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    output_root = (
        Path(args.output_dir).expanduser().resolve()
        if args.output_dir
        else Path(__file__).parents[1] / "exports"
    )
    output_root.mkdir(parents=True, exist_ok=True)

    run_all_pathways(
        workbook_path=workbook_path,
        output_root=output_root,
        max_pathways=args.max_pathways,
        excel_visible=args.excel_visible,
    )


if __name__ == "__main__":
    main()
