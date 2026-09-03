"""xlwings / desktop-Excel recalc engine — native, parallel, resumable.

Ported from the legacy ``src/runner.py`` loop with the throughput work needed to
run hundreds of pathways:

* the 14 MB workbook is parsed **once** with openpyxl to discover the pathway
  list and the output-table ranges; workers reuse that plan;
* Excel runs with calculation on manual, events / screen-updating / alerts off,
  and a single ``app.calculate()`` per pathway;
* identical scenario tuples are computed **once** and their CSVs fanned out to
  the duplicates (FABLE decks routinely repeat a scenario across pathways);
* already-exported pathways are **skipped on restart** (idempotent run dir);
* ``workers > 1`` shards the pathway list across N Excel processes, each with
  its own workbook copy — throughput scales ~linearly until disk/CPU bound.

For thousands of pathways or a licence-free CI, shard at a higher level: run
this engine on M machines / matrix jobs over disjoint ``--pathway-slice`` ranges
and merge the run dirs (they only share CSV files).
"""

from __future__ import annotations

import math
import os
import shutil
import sys
import tempfile
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

from ..deviation import (
    RUN_PATHWAY_COL,
    export_scenario_deviation_summary,
    safe_name,
)
from . import EngineUnavailable, ProgressCallback

TABLE_OUTPUT_SHEETS = [
    "GHG", "PRODUCTION", "TRADE", "JOBS", "FOOD",
    "LAND", "WATER", "N and P", "BIODIVERSITY",
]
DEVIATION_SUMMARY_FILE = "scenario_deviation_summary.csv"


# --------------------------------------------------------------------------
# workbook plan (parsed once, shared with workers)
# --------------------------------------------------------------------------
@dataclass
class Plan:
    sheet_name: str
    selection_col_idx: int
    first_data_row: int
    last_data_row: int
    # (row_number, pathway_name, scenario_tuple)
    pathways: List[Tuple[int, str, Tuple[str, ...]]]
    # {(sheet, table_name): "A1:Z9"}
    table_ranges: Dict[Tuple[str, str], str] = field(default_factory=dict)


def _build_plan(workbook_path: Path) -> Plan:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries

    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    try:
        ws = wb["PATHWAYS selection"]
        if "PathwaysSelection" not in ws.tables:
            raise ValueError("table 'PathwaysSelection' not found on 'PATHWAYS selection'")
        ref = ws.tables["PathwaysSelection"]
        ref = ref if isinstance(ref, str) else ref.ref
        b = range_boundaries(ref)
        min_col, min_row, max_col, max_row = (b[0] or 0), (b[1] or 0), (b[2] or 0), (b[3] or 0)

        headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
        norm = [str(h).strip().lower() if h is not None else "" for h in headers]
        if "selection" not in norm or "pathway" not in norm:
            raise ValueError("PathwaysSelection needs SELECTION and PATHWAY headers")
        sel_idx = min_col + norm.index("selection")
        path_idx = min_col + norm.index("pathway")
        # scenario columns = every column after PATHWAY
        scen_cols = [c for c in range(path_idx + 1, max_col + 1)]

        pathways: List[Tuple[int, str, Tuple[str, ...]]] = []
        for row in range(min_row + 1, max_row + 1):
            name = ws.cell(row, path_idx).value
            if name is None or str(name).strip() == "":
                continue
            scen = tuple(
                "" if ws.cell(row, c).value is None else str(ws.cell(row, c).value).strip()
                for c in scen_cols
            )
            pathways.append((row, str(name).strip(), scen))
        if not pathways:
            raise ValueError("no pathway rows in PathwaysSelection")

        ranges: Dict[Tuple[str, str], str] = {}
        for sheet_name in TABLE_OUTPUT_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            for tname, tobj in wb[sheet_name].tables.items():
                ranges[(sheet_name, tname)] = tobj if isinstance(tobj, str) else tobj.ref

        return Plan(
            sheet_name="PATHWAYS selection",
            selection_col_idx=sel_idx,
            first_data_row=min_row + 1,
            last_data_row=max_row,
            pathways=pathways,
            table_ranges=ranges,
        )
    finally:
        wb.close()


# --------------------------------------------------------------------------
# CSV helpers
# --------------------------------------------------------------------------
def _dedupe_headers(headers: List[object]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for i, h in enumerate(headers, 1):
        base = str(h).strip() if h is not None else f"col_{i}"
        base = base or f"col_{i}"
        n = seen.get(base, 0)
        seen[base] = n + 1
        out.append(base if n == 0 else f"{base}_{n + 1}")
    return out


def _pathway_dir(run_dir: Path, pathway: str) -> Path:
    return run_dir / "tables_per_pathway" / safe_name(pathway)


def _already_done(run_dir: Path, pathway: str, n_tables: int) -> bool:
    d = _pathway_dir(run_dir, pathway)
    return d.is_dir() and len(list(d.glob("*.csv"))) >= n_tables


def _write_table_csv(raw, sheet: str, tname: str, pathway: str, out_file: Path) -> Optional[pd.DataFrame]:
    if not raw:
        return None
    if not isinstance(raw[0], (list, tuple)):
        raw = [raw]
    headers = _dedupe_headers(raw[0] or [])
    df = pd.DataFrame(raw[1:], columns=headers).dropna(how="all")
    if df.empty:
        return None
    col = RUN_PATHWAY_COL if RUN_PATHWAY_COL not in df.columns else f"_{RUN_PATHWAY_COL}"
    df.insert(0, col, pathway)
    out_file.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_file, index=False)
    return df


# --------------------------------------------------------------------------
# one shard = one Excel process over a list of (row, pathway) reps
# --------------------------------------------------------------------------
def _run_shard(
    workbook_str: str,
    run_dir_str: str,
    plan: Plan,
    reps: List[Tuple[int, str]],
    excel_visible: bool,
    _progress: Optional[ProgressCallback] = None,
) -> List[dict]:
    import xlwings as xw
    from openpyxl.utils.cell import get_column_letter

    run_dir = Path(run_dir_str)
    n_tables = len(plan.table_ranges)
    manifest: List[dict] = []

    # private workbook copy so multiple processes never fight over one file handle
    tmp_dir = Path(tempfile.mkdtemp(prefix="fable_wb_"))
    wb_copy = tmp_dir / Path(workbook_str).name
    shutil.copy2(workbook_str, wb_copy)

    app = xw.App(visible=bool(excel_visible), add_book=False)
    com_wb = None
    try:
        app.display_alerts = False
        app.screen_updating = False
        try:
            app.api.EnableEvents = False
        except Exception:
            pass
        try:
            app.calculation = "manual"
        except Exception:
            pass

        com_wb = app.books.open(str(wb_copy.resolve()), update_links=False)
        sheet = com_wb.sheets[plan.sheet_name]
        col_letter = get_column_letter(plan.selection_col_idx)
        clear_range = f"{col_letter}{plan.first_data_row}:{col_letter}{plan.last_data_row}"

        for i, (row_num, pathway) in enumerate(reps, 1):
            if _progress:
                try:
                    _progress(i, len(reps), pathway, "running")
                except Exception:
                    pass
            entry = {"Pathway": pathway, "Row": row_num, "Status": "ok", "Error": "", "Seconds": 0.0}
            start = time.time()

            if _already_done(run_dir, pathway, n_tables):
                entry["Status"] = "ok"
                entry["Error"] = "skipped (already exported)"
                manifest.append(entry)
                continue

            try:
                sheet.range(clear_range).value = None
                sheet.cells(row_num, plan.selection_col_idx).value = "x"
                app.calculate()

                for (sht, tname), ref in plan.table_ranges.items():
                    try:
                        raw = com_wb.sheets[sht].range(ref).value
                    except Exception:
                        continue
                    out_file = _pathway_dir(run_dir, pathway) / f"{safe_name(sht)}__{safe_name(tname)}.csv"
                    _write_table_csv(raw, sht, tname, pathway, out_file)
            except Exception as exc:  # pragma: no cover - COM failure paths
                entry["Status"] = "failed"
                entry["Error"] = str(exc)
            finally:
                entry["Seconds"] = round(time.time() - start, 2)
                manifest.append(entry)
    finally:
        if com_wb is not None:
            try:
                com_wb.close()
            except Exception:
                pass
        try:
            app.quit()
        except Exception:
            pass
        shutil.rmtree(tmp_dir, ignore_errors=True)

    return manifest


def _shards(reps: List[Tuple[int, str]], workers: int) -> List[List[Tuple[int, str]]]:
    workers = max(1, min(workers, len(reps)))
    size = math.ceil(len(reps) / workers)
    return [reps[i : i + size] for i in range(0, len(reps), size)]


def _fanout_duplicates(run_dir: Path, groups: List[List[str]]) -> None:
    for members in groups:
        if len(members) < 2:
            continue
        src = _pathway_dir(run_dir, members[0])
        if not src.is_dir():
            continue
        for dup in members[1:]:
            dst = _pathway_dir(run_dir, dup)
            dst.mkdir(parents=True, exist_ok=True)
            for csv in src.glob("*.csv"):
                df = pd.read_csv(csv)
                pcol = RUN_PATHWAY_COL if RUN_PATHWAY_COL in df.columns else f"_{RUN_PATHWAY_COL}"
                if pcol in df.columns:
                    df[pcol] = dup
                df.to_csv(dst / csv.name, index=False)


def _combine(run_dir: Path, pathways: List[str], plan: Plan) -> None:
    combined_dir = run_dir / "combined_tables"
    combined_dir.mkdir(parents=True, exist_ok=True)
    for (sht, tname) in plan.table_ranges:
        frames: List[pd.DataFrame] = []
        fname = f"{safe_name(sht)}__{safe_name(tname)}.csv"
        for pathway in pathways:
            fp = _pathway_dir(run_dir, pathway) / fname
            if fp.exists():
                try:
                    frames.append(pd.read_csv(fp))
                except Exception:
                    pass
        if frames:
            out = combined_dir / f"{safe_name(sht)}__{safe_name(tname)}__all_pathways.csv"
            pd.concat(frames, ignore_index=True).to_csv(out, index=False)


# --------------------------------------------------------------------------
# engine
# --------------------------------------------------------------------------
class XlwingsEngine:
    name = "xlwings"

    def available(self) -> tuple[bool, str]:
        try:
            import xlwings  # noqa: F401
        except ImportError:
            return False, "xlwings not installed (pip install 'fable-postprocessing[excel]')"
        if sys.platform not in ("win32", "darwin"):
            return False, f"desktop Excel not available on {sys.platform}"
        return True, "xlwings + desktop Excel"

    def recalc_all(
        self,
        workbook_path: Path,
        output_root: Path,
        *,
        max_pathways: Optional[int] = None,
        workers: int = 1,
        pathway_slice: Optional[tuple[int, int]] = None,
        excel_visible: bool = False,
        run_dir: Optional[Path] = None,
        progress_callback: Optional[ProgressCallback] = None,
    ) -> Path:
        ok, why = self.available()
        if not ok:
            raise EngineUnavailable(f"xlwings engine: {why}")

        workbook_path = Path(workbook_path).resolve()
        if not workbook_path.exists():
            raise FileNotFoundError(f"workbook not found: {workbook_path}")

        plan = _build_plan(workbook_path)
        selected = plan.pathways
        if pathway_slice:
            selected = selected[pathway_slice[0] : pathway_slice[1]]
        if max_pathways:
            selected = selected[:max_pathways]
        if not selected:
            raise RuntimeError("no pathways selected")

        output_root = Path(output_root)
        output_root.mkdir(parents=True, exist_ok=True)
        run_dir = run_dir or (
            output_root / f"all_pathways_run_{datetime.now():%Y%m%d_%H%M%S}"
        )
        run_dir.mkdir(parents=True, exist_ok=True)

        # dedupe by scenario tuple -> groups of pathway names
        by_scen: Dict[Tuple[str, ...], List[Tuple[int, str]]] = {}
        for row_num, name, scen in selected:
            by_scen.setdefault(scen, []).append((row_num, name))
        groups = [[name for _, name in members] for members in by_scen.values()]
        reps = [members[0] for members in by_scen.values()]

        print(
            f"Workbook : {workbook_path.name}\n"
            f"Pathways : {len(selected)} selected, {len(reps)} unique scenario tuple(s)\n"
            f"Tables   : {len(plan.table_ranges)}\n"
            f"Workers  : {workers}\n"
            f"Outputs  : {run_dir}",
            flush=True,
        )

        manifest_rows: List[dict] = []
        if workers > 1 and len(reps) > 1:
            shards = _shards(reps, workers)
            with ProcessPoolExecutor(max_workers=len(shards)) as ex:
                futures = {
                    ex.submit(
                        _run_shard, str(workbook_path), str(run_dir), plan, shard,
                        excel_visible, None,
                    ): shard
                    for shard in shards
                }
                done = 0
                for fut in as_completed(futures):
                    rows = fut.result()
                    manifest_rows.extend(rows)
                    done += len(futures[fut])
                    if progress_callback:
                        progress_callback(done, len(reps), "shard complete", "running")
        else:
            manifest_rows = _run_shard(
                str(workbook_path), str(run_dir), plan, reps, excel_visible, progress_callback
            )

        all_names = [name for _, name, _ in selected]
        _fanout_duplicates(run_dir, groups)
        _combine(run_dir, all_names, plan)

        # expand manifest to every pathway (duplicates inherit rep status)
        rep_status = {r["Pathway"]: r for r in manifest_rows}
        full_manifest: List[dict] = []
        for members in groups:
            base = rep_status.get(members[0], {"Status": "ok", "Error": "", "Seconds": 0.0, "Row": 0})
            for name in members:
                row_num = next((rn for rn, n, _ in selected if n == name), 0)
                full_manifest.append({
                    "Pathway": name, "Row": row_num,
                    "Status": base["Status"], "Error": base["Error"],
                    "Seconds": base["Seconds"],
                })
        pd.DataFrame(full_manifest).to_csv(run_dir / "run_manifest.csv", index=False)

        try:
            export_scenario_deviation_summary(
                combined_dir=run_dir / "combined_tables",
                output_csv=run_dir / DEVIATION_SUMMARY_FILE,
            )
        except Exception as exc:  # pragma: no cover
            print(f"warning: deviation summary failed: {exc}", flush=True)

        if progress_callback:
            progress_callback(len(reps), len(reps), "Done", "done")
        print(f"Done -> {run_dir}", flush=True)
        return run_dir
